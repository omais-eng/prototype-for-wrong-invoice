"""
Generate sample invoice PDFs and Excel files for testing AIRP.

Creates:
  - 5 valid invoices (PDF text-based)
  - 3 invalid invoices (price mismatch, missing PO, wrong quantities)
  - 2 duplicate invoices

Output: scripts/sample_invoices/

Usage:
    pip install reportlab openpyxl faker
    python scripts/generate_sample_invoices.py
"""

import os
import random
from datetime import datetime, timedelta

from faker import Faker

fake = Faker()
OUT_DIR = os.path.join(os.path.dirname(__file__), "sample_invoices")
os.makedirs(OUT_DIR, exist_ok=True)

# ── PDF generation (reportlab) ─────────────────────────────────────────────────

def make_pdf_invoice(filename, data):
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch

        path = os.path.join(OUT_DIR, filename)
        doc = SimpleDocTemplate(path, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Header
        elements.append(Paragraph(f"<b>INVOICE</b>", styles["Title"]))
        elements.append(Spacer(1, 0.2 * inch))

        meta = [
            ["Invoice Number:", data["invoice_number"]],
            ["Vendor:", data["vendor_name"]],
            ["Vendor Email:", data["vendor_email"]],
            ["Invoice Date:", data["invoice_date"]],
            ["PO Number:", data.get("po_number", "N/A")],
            ["Currency:", data.get("currency", "USD")],
        ]
        meta_table = Table(meta, colWidths=[2 * inch, 4 * inch])
        meta_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(meta_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Line items
        elements.append(Paragraph("<b>Line Items</b>", styles["Heading2"]))
        item_data = [["Description", "Qty", "Unit Price", "Total"]]
        for item in data["line_items"]:
            item_data.append([
                item["description"],
                str(item["quantity"]),
                f"${item['unit_price']:.2f}",
                f"${item['total']:.2f}",
            ])
        item_table = Table(item_data, colWidths=[3.5 * inch, 1 * inch, 1.25 * inch, 1.25 * inch])
        item_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ]))
        elements.append(item_table)
        elements.append(Spacer(1, 0.2 * inch))

        # Totals
        totals = [
            ["Subtotal:", f"${data['subtotal']:.2f}"],
            ["Tax (8%):", f"${data['tax_amount']:.2f}"],
            ["Total:", f"${data['total_amount']:.2f}"],
        ]
        totals_table = Table(totals, colWidths=[5 * inch, 2 * inch])
        totals_table.setStyle(TableStyle([
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(totals_table)

        doc.build(elements)
        print(f"  ✓ PDF: {filename}")
        return path
    except ImportError:
        # Fallback: write plain text file
        path = os.path.join(OUT_DIR, filename.replace(".pdf", ".txt"))
        with open(path, "w") as f:
            f.write(f"INVOICE\n{'='*40}\n")
            for k, v in data.items():
                if k != "line_items":
                    f.write(f"{k}: {v}\n")
            f.write("\nLINE ITEMS:\n")
            for item in data["line_items"]:
                f.write(f"  {item['description']} | Qty: {item['quantity']} | ${item['unit_price']} | ${item['total']}\n")
            f.write(f"\nSubtotal: ${data['subtotal']}\nTax: ${data['tax_amount']}\nTotal: ${data['total_amount']}\n")
        print(f"  ✓ TXT (reportlab not installed): {os.path.basename(path)}")
        return path


def make_excel_invoice(filename, data):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        path = os.path.join(OUT_DIR, filename)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"

        header_fill = PatternFill("solid", fgColor="2563EB")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        bold = Font(bold=True)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14

        # Title
        ws["A1"] = "INVOICE"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        # Metadata
        meta_rows = [
            ("Invoice Number", data["invoice_number"]),
            ("Vendor Name", data["vendor_name"]),
            ("Vendor Email", data["vendor_email"]),
            ("Invoice Date", data["invoice_date"]),
            ("PO Number", data.get("po_number", "N/A")),
            ("Currency", data.get("currency", "USD")),
        ]
        for i, (label, value) in enumerate(meta_rows, start=3):
            ws.cell(row=i, column=1, value=label).font = bold
            ws.cell(row=i, column=2, value=value)

        # Line items header
        start_row = 11
        headers = ["Description", "Quantity", "Unit Price", "Total"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        for idx, item in enumerate(data["line_items"], start=1):
            row = start_row + idx
            ws.cell(row=row, column=1, value=item["description"]).border = border
            ws.cell(row=row, column=2, value=item["quantity"]).border = border
            ws.cell(row=row, column=3, value=item["unit_price"]).border = border
            ws.cell(row=row, column=4, value=item["total"]).border = border

        # Totals
        total_row = start_row + len(data["line_items"]) + 2
        for label, value in [("Subtotal", data["subtotal"]), ("Tax (8%)", data["tax_amount"]), ("TOTAL", data["total_amount"])]:
            ws.cell(row=total_row, column=3, value=label).font = bold
            ws.cell(row=total_row, column=4, value=value).font = bold
            total_row += 1

        wb.save(path)
        print(f"  ✓ Excel: {filename}")
        return path
    except ImportError:
        print("  ! openpyxl not installed, skipping Excel generation")
        return None


# ── Invoice data templates ─────────────────────────────────────────────────────

def make_invoice_data(invoice_number, vendor_name, vendor_email, po_number,
                      line_items, subtotal, tax_amount, total_amount,
                      date_offset=0):
    invoice_date = (datetime.utcnow() - timedelta(days=date_offset)).strftime("%Y-%m-%d")
    return {
        "invoice_number": invoice_number,
        "vendor_name": vendor_name,
        "vendor_email": vendor_email,
        "invoice_date": invoice_date,
        "po_number": po_number,
        "currency": "USD",
        "line_items": line_items,
        "subtotal": subtotal,
        "tax_amount": tax_amount,
        "total_amount": total_amount,
    }


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f"Generating sample invoices to: {OUT_DIR}\n")

    # 1. Valid Invoice — matches PO-2024-0001
    valid_items_1 = [
        {"description": "Software Consulting Services", "quantity": 10, "unit_price": 150.00, "total": 1500.00},
        {"description": "System Integration Work", "quantity": 5, "unit_price": 200.00, "total": 1000.00},
    ]
    d = make_invoice_data("INV-2024-00001", "Acme Tech Solutions", "billing@acmetech.com",
                          "PO-2024-0001", valid_items_1, 2500.00, 200.00, 2700.00)
    make_pdf_invoice("valid_invoice_001.pdf", d)

    # 2. Valid Invoice — Excel format
    valid_items_2 = [
        {"description": "IT Support Hours", "quantity": 20, "unit_price": 75.00, "total": 1500.00},
        {"description": "Hardware Maintenance", "quantity": 2, "unit_price": 90.00, "total": 180.00},
    ]
    d = make_invoice_data("INV-2024-00002", "CloudBase Systems", "accounts@cloudbase.io",
                          "PO-2024-0002", valid_items_2, 1680.00, 134.40, 1814.40)
    make_excel_invoice("valid_invoice_002.xlsx", d)

    # 3. Valid Invoice PDF
    valid_items_3 = [
        {"description": "Training Workshop (2 days)", "quantity": 1, "unit_price": 2400.00, "total": 2400.00},
    ]
    d = make_invoice_data("INV-2024-00003", "ProTrain Ltd", "invoices@protrain.com",
                          "PO-2024-0003", valid_items_3, 2400.00, 192.00, 2592.00)
    make_pdf_invoice("valid_invoice_003.pdf", d)

    # 4. INVALID — Price mismatch (unit price inflated above contract rate)
    invalid_price = [
        {"description": "Consulting Services", "quantity": 10, "unit_price": 250.00, "total": 2500.00},  # Contract: $150
    ]
    d = make_invoice_data("INV-2024-00010", "Overcharge Corp", "billing@overcharge.com",
                          "PO-2024-0005", invalid_price, 2500.00, 200.00, 2700.00)
    make_pdf_invoice("invalid_price_mismatch.pdf", d)

    # 5. INVALID — Missing PO number
    no_po_items = [
        {"description": "Software License", "quantity": 3, "unit_price": 500.00, "total": 1500.00},
    ]
    d = make_invoice_data("INV-2024-00011", "Mystery Vendor Inc", "billing@mystery.com",
                          "", no_po_items, 1500.00, 120.00, 1620.00)
    make_pdf_invoice("invalid_missing_po.pdf", d)

    # 6. INVALID — Quantity mismatch (PO says 5, invoice says 50)
    wrong_qty = [
        {"description": "Server Racks", "quantity": 50, "unit_price": 1200.00, "total": 60000.00},  # PO qty: 5
    ]
    d = make_invoice_data("INV-2024-00012", "DataCenter Supplies", "ar@dcsupplies.com",
                          "PO-2024-0007", wrong_qty, 60000.00, 4800.00, 64800.00)
    make_pdf_invoice("invalid_quantity_mismatch.pdf", d)

    # 7. DUPLICATE — Same as valid_invoice_001
    make_pdf_invoice("duplicate_of_001.pdf", make_invoice_data(
        "INV-2024-00001",  # Same invoice number
        "Acme Tech Solutions", "billing@acmetech.com",
        "PO-2024-0001", valid_items_1, 2500.00, 200.00, 2700.00
    ))

    # 8. DUPLICATE — Same vendor + amount, slightly different date
    make_pdf_invoice("duplicate_fuzzy.pdf", make_invoice_data(
        "INV-2024-00001-B",  # Different number but same vendor/amount = fuzzy duplicate
        "Acme Tech Solutions", "billing@acmetech.com",
        "PO-2024-0001", valid_items_1, 2500.00, 200.00, 2700.00,
        date_offset=3  # 3 days later
    ))

    # 9. Valid — Excel with multiple line items
    multi_items = [
        {"description": "Development Sprint 1", "quantity": 80, "unit_price": 200.00, "total": 16000.00},
        {"description": "Development Sprint 2", "quantity": 80, "unit_price": 200.00, "total": 16000.00},
        {"description": "Code Review Sessions", "quantity": 10, "unit_price": 150.00, "total": 1500.00},
        {"description": "Documentation", "quantity": 20, "unit_price": 75.00, "total": 1500.00},
    ]
    d = make_invoice_data("INV-2024-00020", "DevShop Partners", "billing@devshop.co",
                          "PO-2024-0010", multi_items, 35000.00, 2800.00, 37800.00)
    make_excel_invoice("valid_invoice_large.xlsx", d)

    # 10. INVALID — Math error (subtotal does not match line items)
    math_error_items = [
        {"description": "Logistics Services", "quantity": 5, "unit_price": 60.00, "total": 300.00},
        {"description": "Warehouse Storage", "quantity": 10, "unit_price": 45.00, "total": 450.00},
    ]
    d = make_invoice_data("INV-2024-00015", "LogiFlow Ltd", "finance@logiflow.com",
                          "PO-2024-0008", math_error_items,
                          999.00,  # Wrong: should be 750.00
                          79.92, 1078.92)
    make_pdf_invoice("invalid_math_error.pdf", d)

    print(f"\n✓ Generated {len(os.listdir(OUT_DIR))} sample invoice files in: {OUT_DIR}")
    print("\nUpload these to AIRP via the dashboard or POST /api/v1/invoices/upload")


if __name__ == "__main__":
    main()
